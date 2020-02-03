"""
Script relativo ao teste de automação do candidato André França
Realiza preenchimento de um formulário do Google.
Os dados estão contidos em uma planilha de excel.


Dependências:
$ pip install openpyxl
$ pip install requests
"""

import requests
from openpyxl import load_workbook

# Leitura da planilha
wb = load_workbook('test.xlsx') # abrindo o Workbook test.xlsx
ws = wb['Planilha1'] # selecionando a planilha 'Planilha1' dentro do Workbook test.xlsx
linha=0

for line in ws:  # iterando em todas as linhas da 'Plan1'
    #print(line[0]) # print a primeira célula da linha
    linha = linha + 1
    #print(linha)

"""
# Utililizando em índices o nome das células como em um dicionário
a1 = ws['A1']
# Imprime o valor da célula C1
print(a1.value)
"""

# Escrita no formulário online
url = 'https://docs.google.com/forms/d/e/1FAIpQLScBRmhzR0d7u7NSnSlTf66E-6NziilDHIo28u8R4gh_5MizXw/formResponse'

for x in range(2,linha+1):
    Observ='Usuário'
    Tarefa='Primeira'
    Item='10'
    Rot= ws['A'+str(x)].value
    Escolha='Opção 1'
    form_data = {'entry.366340186': Observ, 'entry.443668215': Tarefa, 'entry.22026032': Item, 'entry.371113979': Rot, 'entry.489632448': Escolha, 'draftResponse':[], 'pageHistory':0}

    user_agent = {'Referer':url,'User-Agent': "Mozilla/5.0 (X11; Linux i686) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/28.0.1500.52 Safari/537.36"}
    r = requests.post(url, data=form_data, headers=user_agent)
